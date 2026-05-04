import os
import sys
import requests
import asyncio
import imaplib
import email
import subprocess
import pandas as pd

from telethon.sync import TelegramClient
from telethon.sessions import StringSession
from telethon import events

from keep_alive import keep_alive
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill, Alignment, Font


# =========================
# 🔧 PROCESADOR CSV
# =========================
def procesar_csv(input_path, output_path):

    df = pd.read_csv(input_path, dtype=str, sep=",", encoding="latin1")

    cols_eliminar = [
        "ID del bono del jugador","Moneda","Activado Por",
        "Cantidad de Spins","Recuento de Spins",
        "Cantidad de Bono Restante","Bono Wagering",
        "Tipo de Bloqueo","Anulador",
        "ID de transacción","Monto de la Transacción"
    ]

    df = df.drop(columns=[col for col in cols_eliminar if col in df.columns])

    cols_montos = [
        "Importe canjeado","Importe",
        "Requisitos de apuesta restantes",
        "Monto de facturación",
        "Cantidad de Bono Restante"
    ]

    for col in cols_montos:
        if col in df.columns:
            df[col] = (
                df[col].astype(str)
                .str.replace(".", "", regex=False)
                .str.replace(",", "", regex=False)
            )

            df[col] = pd.to_numeric(df[col], errors="coerce")

            def corregir(x):
                if pd.isna(x): return x
                if x >= 1_000_000_000: return x / 100_000_000
                elif x >= 10000: return x / 10
                return x

            df[col] = df[col].apply(corregir)

    # detectar fecha
    col_fecha = next(
        (c for c in df.columns if "fecha" in c.lower() and "activ" in c.lower()),
        None
    )

    if not col_fecha:
        raise Exception("No se encontró columna fecha")

    df["Fecha_limpia"] = pd.to_datetime(df[col_fecha], errors="coerce").dt.date

    # duplicados
    df["duplicado"] = df.duplicated(subset=["ID del jugador"], keep=False)
    df = df[df["duplicado"]]

    df["mismo_dia"] = df.duplicated(
        subset=["ID del jugador", "Fecha_limpia"], keep=False
    )

    df = df.sort_values(by=["mismo_dia","ID del jugador"], ascending=[False,True])

    # exportar
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Datos")
        ws = writer.sheets["Datos"]

        fill_rojo = PatternFill("solid", fgColor="FF0000")
        fill_amarillo = PatternFill("solid", fgColor="FFFF00")
        align = Alignment(horizontal="center", vertical="center")

        col_importe = df.columns.get_loc("Importe canjeado")+1
        col_mismo = df.columns.get_loc("mismo_dia")+1

        for row in range(2, len(df)+2):
            val = ws.cell(row, col_importe).value or 0
            mismo = ws.cell(row, col_mismo).value

            for col in range(1, len(df.columns)+1):
                ws.cell(row, col).alignment = align

            if mismo:
                for col in range(1, len(df.columns)+1):
                    ws.cell(row, col).fill = fill_amarillo
            elif float(val) > 10000:
                for col in range(1, len(df.columns)+1):
                    ws.cell(row, col).fill = fill_rojo

        ws.cell(row=len(df)+4, column=1,
                value="Generado por Erik Sanzana").font = Font(bold=True)

    return output_path


# =========================
# 🚀 INIT
# =========================
keep_alive()
sys.stdout.reconfigure(encoding='utf-8')

string_session = os.getenv('STRING_SESSION')
api_id = os.getenv('API_ID')
api_hash = os.getenv('API_HASH')

group_id_to_monitor1 = int(os.getenv('GROUP_ID_TO_MONITOR1'))
group_id_to_monitor2 = int(os.getenv('GROUP_ID_TO_MONITOR2'))
group_id_to_monitor3 = int(os.getenv('GROUP_ID_TO_MONITOR3'))
group_id_to_forward = int(os.getenv('GROUP_ID_TO_FORWARD'))

gmail_user = os.getenv("GMAIL_USER")
gmail_pass = os.getenv("GMAIL_PASS")

client = TelegramClient(StringSession(string_session), api_id, api_hash)

allowed_groups = [
    group_id_to_forward,
    group_id_to_monitor3
]


# =========================
# 📊 STATUS SERVICIOS
# =========================



def get_astropay_status():
    try:
        response = requests.get(
            "https://status.astropay.com/api/v2/status.json", timeout=5)

        data = response.json()
        description = data["status"]["description"]
        indicator = data["status"]["indicator"]

        emoji = {
            "none": "🟢",
            "minor": "🟡",
            "major": "🔴",
            "critical": "❌"
        }.get(indicator, "❓")

        return f"{emoji} *AstroPay*: {description}"
    except Exception as e:
        return f"⚠️ *AstroPay*: Error ({e})"


def get_kushki_status():
    try:
        response = requests.get(
            "https://status.kushkipagos.com/api/v2/status.json", timeout=5)

        data = response.json()
        description = data["status"]["description"]
        indicator = data["status"]["indicator"]

        emoji = {
            "none": "🟢",
            "minor": "🟡",
            "major": "🔴",
            "critical": "❌"
        }.get(indicator, "❓")

        return f"{emoji} *Kushki*: {description}"
    except Exception as e:
        return f"⚠️ *Kushki*: Error ({e})"


def get_transbank_status():
    try:
        response = requests.get(
            "https://status.transbankdevelopers.cl/api/v2/status.json", timeout=5)

        data = response.json()
        description = data["status"]["description"]
        indicator = data["status"]["indicator"]

        emoji = {
            "none": "🟢",
            "minor": "🟡",
            "major": "🔴",
            "critical": "❌"
        }.get(indicator, "❓")

        return f"{emoji} *Transbank*: {description}"
    except Exception as e:
        return f"⚠️ *Transbank*: Error ({e})"


def get_skinsback_status():
    try:
        response = requests.get("https://skinsback.com", timeout=5)
        return "🟢 *Skinsback*: Activo" if response.status_code == 200 else f"🔴 HTTP {response.status_code}"
    except Exception as e:
        return f"⚠️ *Skinsback*: Error ({e})"


def get_coinpaid_status():
    try:
        response = requests.get(
            "https://app.cryptoprocessing.com/api/v2/ping", timeout=5)

        return "🟢 *CoinPaid*: Activo" if response.status_code == 200 else f"🔴 HTTP {response.status_code}"
    except Exception as e:
        return f"⚠️ *CoinPaid*: Error ({e})"


@client.on(events.NewMessage(pattern=r'^/servicios$', chats=allowed_groups))
async def check_services_status(event):

    statuses = [
        get_astropay_status(),
        get_kushki_status(),
        get_transbank_status(),
        get_skinsback_status(),
        get_coinpaid_status()
    ]

    message = "**Estado actual de servicios:**\n\n" + "\n".join(statuses)

    await client.send_message(event.chat_id, message, parse_mode='Markdown')


# =========================
# 🔍 SHERLOCK
# =========================
def buscar_usuario_con_sherlock(nick):
    try:
        result = subprocess.run(
            ["sherlock", nick],
            capture_output=True, text=True, timeout=500
        )
        return result.stdout
    except Exception as e:
        return str(e)


@client.on(events.NewMessage(pattern=r'^/nick\s+(.+)', chats=[group_id_to_forward]))
async def handler_sherlock(event):

    nick = event.pattern_match.group(1).strip()
    await event.respond("🔍 Buscando...")

    resultado = buscar_usuario_con_sherlock(nick)

    await event.respond(f"```{resultado}```", parse_mode="Markdown")


# =========================
# 📧 EMAIL
# =========================
def extraer_cuerpo_email(msg):
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                return part.get_payload(decode=True).decode(errors="ignore")
            if part.get_content_type() == "text/html":
                html = part.get_payload(decode=True).decode(errors="ignore")
                return BeautifulSoup(html, "lxml").get_text()
    return msg.get_payload(decode=True).decode(errors="ignore")


async def revisar_correos_gmail():
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(gmail_user, gmail_pass)
        mail.select("inbox")

        status, mensajes = mail.search(None, '(UNSEEN)')

        for num in mensajes[0].split():
            status, data = mail.fetch(num, "(RFC822)")
            msg = email.message_from_bytes(data[0][1])

            cuerpo = extraer_cuerpo_email(msg)

            await client.send_message(
                group_id_to_forward,
                f"📩 Nuevo correo:\n\n{cuerpo[:500]}"
            )

            mail.store(num, '+FLAGS', '\\Seen')

        mail.logout()

    except Exception as e:
        print("Error Gmail:", e)


async def loop_correos():
    while True:
        await revisar_correos_gmail()
        await asyncio.sleep(15)


# =========================
# 📂 CSV HANDLER (🔥 NUEVO)
# =========================
@client.on(events.NewMessage(func=lambda e: e.file and e.file.name.endswith(".csv")))
async def handle_csv(event):

    try:
        await event.respond("📥 Procesando CSV...")

        input_file = f"input_{event.id}.csv"
        output_file = f"output_{event.id}.xlsx"

        await event.download_media(file=input_file)

        procesar_csv(input_file, output_file)

        await client.send_file(
            event.chat_id,
            output_file,
            caption="✅ CSV procesado correctamente"
        )

        os.remove(input_file)
        os.remove(output_file)

    except Exception as e:
        await event.respond(f"❌ Error:\n{e}")


# =========================
# ▶️ MAIN
# =========================
async def main():
    await client.start()
    print("Bot activo 🚀")

    client.loop.create_task(loop_correos())

    await client.run_until_disconnected()


with client:
    client.loop.run_until_complete(main())
